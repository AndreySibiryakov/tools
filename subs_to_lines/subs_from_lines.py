#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import pickle
import sys
import os
import re
import transcript_mapping as tr
import copy_tags_to_langs as cp
import csv

doc_path = 'c:/SVN/content/docs/m3/Scenario_docs/Metro3 Lines_dlc_2.xlsx'
doc_path_us = 'c:/SVN/content/docs/m3/Scenario_docs/Metro3 Lines_dlc_2.xlsx'
data_path = 'c:/SVN/content/docs/m3/Scenario_docs/m3_lines_data.pckl'
csv_file_path = 'd:/downloads/VO_LocalizationExport_17052019.csv'
level_sounds_data = {
    'm3_dlc2': 'c:/SVN/content/sounds.{lang}/voices/m3_dlc2/'}

us = 'us'
ru = 'ru'
uk = 'uk'
it = 'it'
fr = 'fr'
de = 'de'
es = 'es'
jp = 'jp'
level = 'level'

# Langs present in excel
langs = [us, uk, ru]
# Langs from localize direct. Exported to .csv
csv_langs = [it, fr, de, es, jp]

us_subs_col = 'S'
ru_subs_col = 'T'
uk_subs_col = 'U'
intensity_col = 'Q'
# Other langs not in excel, in LD only
name_col = 'C'
level_col = 'B'
sheet = 'Metro 3 Lines'
data = {}
name_col_vals = []
level_col_vals = []
us_col_vals = []
ru_col_vals = []
uk_col_vals = []

neutral_tag = r'<ffx:neutral>'
txt_ext = '.txt'

# Taken from csv file
phrase_id = "String Identifier"
it_subs_col = "Italian text"
fr_subs_col = "French text"
de_subs_col = "German text"
es_subs_col = "Spanish text"
jp_subs_col = "Japanese text"

lang_col_map = {us: us_subs_col,
                ru: ru_subs_col,
                uk: uk_subs_col,
                it: it_subs_col,
                fr: fr_subs_col,
                de: de_subs_col,
                es: es_subs_col,
                jp: jp_subs_col}


def load(v):
    with open(v, 'r') as p:
        data = pickle.loads(p.read())
    return data


def save_txt(path, data):
    with open(path, 'w+') as p:
        p.write(data.encode("UTF-8"))


def save(v, data):
    with open(v, 'w+') as p:
        pickle.dump(data, p)


def start():
    print '\n' + 'Enter command:'
    print 'update, batch, find, exit' + '\n'
    command = raw_input()
    if command == 'update':
        print 'Will take near 2 minutes. Updating...' + '\n'
        update()
        start()
    elif command == 'batch':
        batch()
        start()
    elif command == 'find':
        find()
        start()
    elif command == 'exit':
        sys.exit()
    else:
        print 'Command is not recognized.', command
        start()


def check_data():
    global data

    if data:
        return
    elif not os.path.exists(data_path):
        print 'Data is not collected yet. Collecting...'
        update()
    else:
        data = load(data_path)


def transliterate_ru(val):
    tr_val = ''

    for v in val:
        tr_val += tr.ru_lang_mapping.get(v, '')

    return tr_val


def transliterate_uk(val):
    tr_val = ''

    for v in val:
        tr_val += tr.uk_lang_mapping.get(v, '')

    return tr_val


def transliterate(val, lang):
    if lang == ru:
        return transliterate_ru(val)
    elif lang == uk:
        return transliterate_uk(val)
    else:
        return val


def find():
    print 'Searches in "us" subs only.'
    print 'Input is converted to lower case.'
    text = raw_input()
    found = False
    check_data()

    for name, name_data in data.iteritems():
        if not name_data[us]:
            continue

        # Deletes <facefx tags> from the text
        sub = re.sub('\<.*\>', '', name_data[us].lower())
        if text.lower() in sub:
            found = True
            print '\n', name
            print name_data[us].encode("ascii", 'replace') + '\n\n'

    if not found:
        print 'No matches found for', text


def update():
    global data
    us_data = update_us()
    wb = load_workbook(doc_path)
    sh = wb[sheet]
    name_col_data = [cell.value for cell in sh[name_col]]
    level_col_data = [cell.value for cell in sh[level_col]]
    # us_subs_col = [cell.value for cell in sh[us_subs_col]]
    ru_subs_col_data = [cell.value for cell in sh[ru_subs_col]]
    uk_subs_col_data = [cell.value for cell in sh[uk_subs_col]]

    for ind, name in enumerate(name_col_data):
        name_data = us_data.get(name, {us: ''})
        name_data[ru] = ru_subs_col_data[ind]
        name_data[uk] = uk_subs_col_data[ind]
        name_data[level] = level_col_data[ind]
        data[name] = name_data

    save(data_path, data)


def update_us():
    # Seperated us as tagged and stored local because of conflicts
    wb_us = load_workbook(doc_path_us)
    sh_us = wb_us[sheet]
    us_subs_col_data = [cell.value for cell in sh_us[us_subs_col]]
    name_col_data = [cell.value for cell in sh_us[name_col]]
    data = {name: {us: sub}
            for name, sub in zip(name_col_data, us_subs_col_data)}
    return data


def prepare_text():
    pass


def get_level_and_name(val):
    data = val.split('@')

    if len(data) != 3:
        return None
    else:
        return data[1], data[2]


def batch():
    check_data()

    for name, name_data in data.iteritems():
        if name_data[level] not in level_sounds_data.keys():
            continue

        us_tagged_text = name_data[us]
        # Empty cell
        if not us_tagged_text:
            continue
        us_tags_data = cp.get_tags_data(us_tagged_text)

        for lang in langs:
            # Enclose code below with try
            try:
                sub_path = level_sounds_data.get(name_data[level]).format(
                    lang=lang) + name + txt_ext
                if lang == us:
                    sub_data = us_tagged_text + ' ' + neutral_tag
                else:
                    sub_data = name_data[lang]
                    sub_data = transliterate(sub_data.lower(), lang)
                    # Copies tags from us for other langs
                    sub_data = cp.add_tags_data(
                        us_tags_data, sub_data) + ' ' + neutral_tag
            except:
                print 'Error with', name
                continue

            save_txt(sub_path, sub_data)

    # For seperate csv file
    with open(csv_file_path, mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)

        for row in csv_reader:
            level_name_data = get_level_and_name(row[phrase_id])
            # Not spoken voice row
            # Spoken structure is:
            # "voices@m3_02_dead_moscow@dmc_000_5_spartans_radio_01_02"
            if not level_name_data:
                continue
            else:
                l, name = level_name_data

            if l not in level_sounds_data.keys():
                continue

            us_tagged_text = data.get(name, {}).get(us)

            if not us_tagged_text:
                # Extra line not present in original excel file
                continue
            # us_tagged_text = data[name][us]

            if not us_tagged_text:
                continue
            us_tags_data = cp.get_tags_data(us_tagged_text)

            for lang in csv_langs:
                try:
                    sub_data = row[lang_col_map[lang]]
                except:
                    # No lang in .csv file
                    continue
                sub_data = unicode(sub_data, "utf-8")
                # If translation not done yet
                if not sub_data:
                    continue
                # Copies tags from us for other langs
                sub_data = cp.add_tags_data(
                    us_tags_data, sub_data) + ' ' + neutral_tag

                sub_path = level_sounds_data.get(l).format(
                    lang=lang) + name + txt_ext

                if not os.path.exists(os.path.dirname(sub_path)):
                    os.makedirs(os.path.dirname(sub_path))

                save_txt(sub_path, sub_data)


start()
'''
check_data()
word = 'sir'
us_col = data[us_subs_col]
levels = data[level_col]
us_col = [v for v in us_col if v]
thanks = [u for l, u in zip(levels, us_col)
          if word in u.lower() and l == 'm3_dlc1']
dlc_len = len([l for l in levels if l == 'm3_dlc1'])
print 'Word "', word, '" used ', len(thanks), 'times in ', dlc_len, 'lines'
print 'In every', int(dlc_len/len(thanks)), '-th phrase'
# print thanks

check_data()

import pprint

print data['dlc1_74_2_quarrel_01_14'][ru]
print data['dlc1_74_2_quarrel_01_14'][us]
print data['dlc1_74_2_quarrel_01_14'][uk]
'''
